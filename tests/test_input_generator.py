from pathlib import Path

import openpyxl

from overlap_calculator.services.input_generator import (
    generate_input_json,
    generate_inputs,
    generate_inputs_with_skips,
)


def test_generate_inputs_reads_only_tddft_out_files(tmp_path: Path) -> None:
    files_dir = tmp_path / "files"
    files_dir.mkdir()
    one = files_dir / "mol1.out"
    two = files_dir / "mol2.out"
    opt = files_dir / "opt.out"
    one.write_text("# td=(nstates=10) b3pw91/6-31g(d,p)\n", encoding="utf-8")
    two.write_text(
        " Excited State   1:   Singlet-A    1.8647 eV  664.91 nm  f=0.3040\n",
        encoding="utf-8",
    )
    opt.write_text("# opt b3lyp/6-31g(d)\n", encoding="utf-8")

    items = generate_inputs(files_dir)

    assert len(items) == 2
    assert {item.sample_id for item in items} == {"mol1", "mol2"}
    assert {Path(item.source_path).name for item in items} == {"mol1.out", "mol2.out"}


def test_generate_input_json_writes_source_paths(tmp_path: Path) -> None:
    files_dir = tmp_path / "files"
    files_dir.mkdir()
    (files_dir / "sample.out").write_text("# td=(nstates=10)\n", encoding="utf-8")
    (files_dir / "opt.out").write_text("# opt b3lyp/6-31g(d)\n", encoding="utf-8")
    out = tmp_path / "input.json"

    items = generate_input_json(files_dir, out)

    assert len(items) == 1
    text = out.read_text(encoding="utf-8")
    assert '"sample_id": "sample"' in text
    assert '"source_path": "files/sample.out"' in text
    assert '"input_type": "theoretical"' in text


def _write_two_sheet_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    first = workbook.active
    first.title = "default"
    first.append(["wavelength_nm", "sample_default"])
    for row in range(10):
        first.append([300.0 + row * 10.0, 0.10 + row * 0.01])
    raw = workbook.create_sheet("raw")
    raw.append(["wavelength_nm", "sample_raw"])
    for row in range(10):
        raw.append([300.0 + row * 10.0, 0.50 + row * 0.02])
    workbook.save(path)


def test_generate_inputs_defaults_to_first_sheet(tmp_path: Path) -> None:
    files_dir = tmp_path / "files"
    files_dir.mkdir()
    workbook_path = files_dir / "dye_panel.xlsx"
    _write_two_sheet_workbook(workbook_path)

    items, _skipped = generate_inputs_with_skips(files_dir)

    assert len(items) == 1
    assert items[0].input_type == "experimental"
    assert items[0].series_name == "sample_default"
    assert items[0].sheet_name is None


def test_generate_inputs_honors_sheet_overrides(tmp_path: Path) -> None:
    files_dir = tmp_path / "files"
    files_dir.mkdir()
    workbook_path = files_dir / "dye_panel.xlsx"
    _write_two_sheet_workbook(workbook_path)

    items, _skipped = generate_inputs_with_skips(
        files_dir,
        sheet_overrides={workbook_path: "raw"},
    )

    assert len(items) == 1
    assert items[0].series_name == "sample_raw"
    assert items[0].sheet_name == "raw"
    assert items[0].sample_id == "sample_raw"


def test_generate_inputs_falls_back_to_later_experimental_sheet(tmp_path: Path) -> None:
    files_dir = tmp_path / "files"
    files_dir.mkdir()
    workbook_path = files_dir / "photochemcad.xlsx"

    workbook = openpyxl.Workbook()
    readme = workbook.active
    readme.title = "README"
    readme.append(["Organic UV-Vis spreadsheet dataset", "notes"])
    readme.append(["not a spectrum", "metadata"])
    raw = workbook.create_sheet("program_input_absorbance")
    raw.append(["wavelength_nm", "Coumarin_1", "Pyrene"])
    for row in range(10):
        raw.append([219.0 + row, 0.10 + row * 0.01, 0.20 + row * 0.01])
    workbook.save(workbook_path)

    items, skipped = generate_inputs_with_skips(files_dir)

    assert not skipped
    assert {item.series_name for item in items} == {"Coumarin_1", "Pyrene"}
    assert {item.sheet_name for item in items} == {"program_input_absorbance"}


def test_generate_inputs_uses_chk_name_when_present(tmp_path: Path) -> None:
    files_dir = tmp_path / "files"
    files_dir.mkdir()
    (files_dir / "slurm-5473089.out").write_text(
        "%chk=B1_td.chk\n# td=(nstates=10) b3pw91/6-31g(d,p)\n",
        encoding="utf-8",
    )
    (files_dir / "slurm-5473092.out").write_text(
        "# td=(nstates=10)\n"
        " Excited State   1:   Singlet-A    1.8647 eV  664.91 nm  f=0.3040\n",
        encoding="utf-8",
    )

    items = generate_inputs(files_dir)

    by_source = {Path(item.source_path).name: item.sample_id for item in items}
    assert by_source["slurm-5473089.out"] == "B1_td"
    assert by_source["slurm-5473092.out"] == "slurm-5473092"
